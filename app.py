import streamlit as st
import requests
import pandas as pd
import plotly.graph_objects as go
import random, re, json, time, io, hashlib
import xml.etree.ElementTree as ET
from deep_translator import GoogleTranslator
from datetime import datetime
from urllib.parse import quote_plus
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ════════════════════════════════════════════════════════════════════════════
# 🔐 CONFIGURACIÓN DE CONTRASEÑAS — Modifica aquí
# ════════════════════════════════════════════════════════════════════════════
USUARIOS = {
    "admin":      "860059G",
}

# ════════════════════════════════════════════════════════════════════════════
# 🎨 ESTILOS GLOBALES
# ════════════════════════════════════════════════════════════════════════════
CSS = """
<style>
    /* Fondo oscuro global */
    .stApp { background-color: #0f1923; color: #e2e8f0; font-family: 'Segoe UI', sans-serif; }
    
    /* Sidebar oscuro */
    .css-1d391d5 { background-color: #0a1118 !important; }
    section[data-testid="stSidebar"] { background-color: #0a1118 !important; }
    section[data-testid="stSidebar"] .stMarkdown { color: #cbd5e1; }

    /* Título principal */
    h1 { color: #38bdf8 !important; text-align: center; font-size: 2rem !important; margin-bottom: 4px !important; }
    h2 { color: #7dd3fc !important; margin-top: 18px !important; }
    h3 { color: #bae6fd !important; }

    /* Inputs */
    .stTextInput > div > div > input,
    .stTextarea > div > div > textarea {
        background-color: #162231 !important;
        color: #f1f5f9 !important;
        border: 1px solid #2d4a6f !important;
        border-radius: 8px !important;
    }
    .stTextInput > div > div > input:focus,
    .stTextarea > div > div > textarea:focus {
        border-color: #38bdf8 !important;
        box-shadow: 0 0 0 2px rgba(56,189,248,0.3) !important;
    }
    .stTextInput label, .stTextarea label { color: #7dd3fc !important; font-weight: 600; }

    /* Selectbox / Multiselect */
    .stSelectbox > div > div, .stMultiselect > div > div {
        background-color: #162231 !important;
        color: #f1f5f9 !important;
        border: 1px solid #2d4a6f !important;
        border-radius: 8px !important;
    }
    .stSelectbox label, .stMultiselect label { color: #7dd3fc !important; font-weight: 600; }

    /* Botones */
    .stButton > button {
        background: linear-gradient(135deg, #1e6fa0, #38bdf8) !important;
        color: #fff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 12px 28px !important;
        font-size: 1.05rem !important;
        font-weight: 700 !important;
        cursor: pointer;
        transition: all 0.25s ease;
        width: 100%;
    }
    .stButton > button:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(56,189,248,0.35) !important; }

    /* Cards / containers */
    .card {
        background: #162231;
        border: 1px solid #2d4a6f;
        border-radius: 14px;
        padding: 22px;
        margin-bottom: 16px;
    }

    /* Login box */
    .login-box {
        background: linear-gradient(145deg, #0f1e2e, #162231);
        border: 1px solid #2d4a6f;
        border-radius: 18px;
        padding: 44px 36px;
        max-width: 420px;
        margin: 60px auto;
        box-shadow: 0 8px 32px rgba(0,0,0,0.4);
    }
    .login-box h2 { text-align: center; color: #38bdf8; margin-bottom: 8px; }
    .login-box p  { text-align: center; color: #64748b; margin-bottom: 24px; }

    /* Badges */
    .badge-ambas {
        background: linear-gradient(135deg, #166534, #22c55e);
        color: #fff; padding: 3px 10px; border-radius: 12px;
        font-size: 0.75rem; font-weight: 700; display: inline-block;
    }
    .badge-una {
        background: linear-gradient(135deg, #b45309, #f59e0b);
        color: #fff; padding: 3px 10px; border-radius: 12px;
        font-size: 0.75rem; font-weight: 700; display: inline-block;
    }
    .badge-tesis {
        background: linear-gradient(135deg, #6d28d9, #8b5cf6);
        color: #fff; padding: 3px 10px; border-radius: 12px;
        font-size: 0.72rem; font-weight: 700; display: inline-block;
    }
    .badge-articulo {
        background: linear-gradient(135deg, #0c4a6e, #0ea5e9);
        color: #fff; padding: 3px 10px; border-radius: 12px;
        font-size: 0.72rem; font-weight: 700; display: inline-block;
    }

    /* Result card */
    .result-card {
        background: #1a2d3d;
        border: 1px solid #2d4a6f;
        border-radius: 12px;
        padding: 16px 20px;
        margin-bottom: 10px;
        transition: border-color 0.2s;
    }
    .result-card:hover { border-color: #38bdf8; }

    /* Stats row */
    .stats-row { display: flex; gap: 12px; flex-wrap: wrap; justify-content: center; margin-bottom: 20px; }
    .stat-box {
        background: #162231; border: 1px solid #2d4a6f; border-radius: 12px;
        padding: 12px 20px; text-align: center; flex: 1; min-width: 110px;
    }
    .stat-box .num { font-size: 1.7rem; font-weight: 800; color: #38bdf8; }
    .stat-box .lbl { font-size: 0.78rem; color: #64748b; margin-top: 2px; }

    /* Divider */
    hr { border-color: #2d4a6f !important; }

    /* Checkbox */
    .stCheckbox label { color: #cbd5e1 !important; }

    /* Expander */
    .streamlit-expander { border-color: #2d4a6f !important; background: #162231 !important; }
    .streamlit-expander p { color: #cbd5e1; }

    /* scrollbar */
    ::-webkit-scrollbar { width: 6px; }
    ::-webkit-scrollbar-track { background: #0f1923; }
    ::-webkit-scrollbar-thumb { background: #2d4a6f; border-radius: 3px; }

    .centered { text-align: center; }
    .subtle { color: #64748b; font-size: 0.85rem; }

    /* Tags multiselect pills */
    .stMultiselect span.css-1n7v3ny-multiwordOption { background-color: #1e3a5f !important; color: #7dd3fc !important; }
</style>
"""

# ════════════════════════════════════════════════════════════════════════════
# ⚙️ CONSTANTES Y HELPERS
# ════════════════════════════════════════════════════════════════════════════
HEADERS = {
    'User-Agent': 'AcademicSearchBot/2.0 (Research; mailto:investigador@universidad.edu)',
    'Accept': 'application/xml, application/json, text/html'
}
IDIOMAS_LABEL = {'en': '🇬🇧 Inglés', 'es': '🇪🇸 Español', 'pt': '🇧🇷 Portugués'}

REPOSITORIOS_OAIPMH = {
    "🇵🇪 UNMSM (San Marcos)":  "https://cybertesis.unmsm.edu.pe/oai/request",
    "🇵🇪 UNI":                  "https://repositorio.uni.edu.pe/oai/request",
    "🇵🇪 PUCP":                 "https://repositorio.pucp.edu.pe/index/oai/request",
    "🇵🇪 UCV (César Vallejo)":  "https://repositorio.ucv.edu.pe/oai/request",
    "🇵🇪 UNSA (Arequipa)":      "https://repositorio.unsa.edu.pe/oai/request",
    "🇵🇪 UNT (Trujillo)":       "https://dspace.unitru.edu.pe/oai/request",
    "🇵🇪 UNCP (Centro)":        "https://repositorio.uncp.edu.pe/oai/request",
    "🇨🇴 U. Nacional":          "https://repositorio.unal.edu.co/oai/request",
    "🇨🇴 U. Antioquia":         "https://bibliotecadigital.udea.edu.co/oai/request",
    "🇨🇴 U. Rosario":           "https://repository.urosario.edu.co/oai/request",
    "🇲🇽 UNAM (RU-DGTIC)":      "https://ru.dgb.unam.mx/oai/request",
    "🇲🇽 IPN":                   "https://tesis.ipn.mx/oai/request",
    "🇲🇽 Colmex":                "https://repositorio.colmex.mx/oai/request",
    "🇨🇱 U. Chile":              "https://repositorio.uchile.cl/oai/request",
    "🇨🇱 USACH":                 "https://repositorio.usach.cl/oai/request",
    "🇪🇨 PUCE":                  "http://repositorio.puce.edu.ec/oai/request",
    "🇪🇨 UPS (Salesiana)":       "https://dspace.ups.edu.ec/oai/request",
    "🇪🇨 ESPE":                  "http://repositorio.espe.edu.ec/oai/request",
    "🇨🇷 TEC":                   "https://repositoriotec.tec.ac.cr/oai/request",
    "🇨🇷 UCR":                   "https://www.kerwa.ucr.ac.cr/oai/request",
}

FUENTES_LISTA = ["OpenAlex","CrossRef","PubMed","Semantic Scholar","Europe PMC","DOAJ","arXiv","SciELO"]

# ─── Utilidades ─────────────────────────────────────────────────────────────
def normalizar(texto):
    if not texto: return ""
    texto = texto.lower()
    for o, r in {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n','ç':'c','ü':'u'}.items():
        texto = texto.replace(o, r)
    return texto

_cache_trad = {}
def traducir(texto, destino):
    key = f"{texto}_{destino}"
    if key in _cache_trad: return _cache_trad[key]
    try:
        res = GoogleTranslator(source='es', target=destino).translate(texto)
        _cache_trad[key] = res
        return res
    except:
        return texto

def variable_en_texto(texto, variable):
    if not variable: return True
    texto_n, var_n = normalizar(texto), normalizar(variable)
    if var_n in texto_n: return True
    palabras = [p for p in var_n.split() if len(p) > 3]
    return len(palabras) >= 2 and all(p in texto_n for p in palabras)

def clasificar_relevancia(texto, var1, var2):
    if not var2:
        return 'ambas' if variable_en_texto(texto, var1) else 'ninguna'
    var1_en = traducir(var1, 'en').lower() if var1 else ''
    var2_en = traducir(var2, 'en').lower() if var2 else ''
    tiene_v1 = variable_en_texto(texto, var1) or (var1_en and variable_en_texto(texto, var1_en))
    tiene_v2 = variable_en_texto(texto, var2) or (var2_en and variable_en_texto(texto, var2_en))
    if tiene_v1 and tiene_v2: return 'ambas'
    if tiene_v1 or tiene_v2:  return 'una'
    return 'ninguna'

def parsear_tema(tema):
    tema_lower = tema.lower()
    pais, tema_sin_pais = None, tema
    match = re.search(r'\s+en\s+([a-záéíóúüñ\s]+)$', tema_lower)
    if match:
        pais = match.group(1).strip().title()
        tema_sin_pais = re.sub(r'\s+en\s+([a-záéíóúüñ\s]+)$', '', tema, flags=re.IGNORECASE).strip()
    var1, var2 = tema_sin_pais, None
    for sep in [' y ', ' and ', ' e ']:
        if sep in tema_sin_pais.lower():
            idx = tema_sin_pais.lower().find(sep)
            var1, var2 = tema_sin_pais[:idx].strip(), tema_sin_pais[idx+len(sep):].strip()
            break
    return var1, var2, pais

def extraer_apellido(nombre):
    if not nombre: return ""
    nombre = nombre.strip()
    if ',' in nombre: return nombre.split(',')[0].strip()
    partes = nombre.split()
    if len(partes) >= 2:
        if len(partes) >= 3 and partes[-2].lower() in ['de','del','la','los','van','von','da','dos','das']:
            return f"{partes[-2]} {partes[-1]}"
        return partes[-1]
    return nombre

def formatear_cita_apa(autores_lista, año):
    if not autores_lista: return f"({año})"
    apellidos = [a for a in (extraer_apellido(x) for x in autores_lista) if a]
    if not apellidos:        return f"({año})"
    if len(apellidos) == 1:  return f"{apellidos[0]} ({año})"
    if len(apellidos) == 2:  return f"{apellidos[0]} y {apellidos[1]} ({año})"
    return f"{apellidos[0]} et al. ({año})"

def autores_display(autores_lista, max_show=3):
    if not autores_lista: return ""
    d = "; ".join(autores_lista[:max_show])
    if len(autores_lista) > max_show: d += " et al."
    return d

def safe_int(val, default=0):
    try: return int(val or default)
    except: return default

def get_abstract_openalex(work):
    try:
        inv = work.get("abstract_inverted_index", {})
        if inv:
            palabras = {}
            for p, pos in inv.items():
                for i in pos: palabras[i] = p
            return ' '.join([palabras[i] for i in sorted(palabras.keys())])
    except: pass
    return ""

# ════════════════════════════════════════════════════════════════════════════
# 🏛️ OAI-PMH
# ════════════════════════════════════════════════════════════════════════════
def cosecha_oaipmh(base_url, nombre_repo, var1, var2, año_inicio, año_fin, max_paginas=2):
    resultados = []
    params = {
        'verb': 'ListRecords', 'metadataPrefix': 'oai_dc',
        'from': f'{año_inicio}-01-01', 'until': f'{año_fin}-12-31'
    }
    pagina = 0
    años_busqueda = list(range(año_inicio, año_fin+1))
    while pagina < max_paginas:
        try:
            r = requests.get(base_url, params=params, timeout=20, headers=HEADERS)
            if r.status_code != 200: break
            root = ET.fromstring(r.content)
            error = root.find('.//{http://www.openarchives.org/OAI/2.0/}error')
            if error is not None: break
            records = root.findall('.//{http://www.openarchives.org/OAI/2.0/}record')
            if not records: break
            for record in records:
                try:
                    dc = record.find('.//{http://www.openarchives.org/OAI/2.0/oai_dc/}dc')
                    if dc is None: continue
                    titulos   = [t.text for t in dc.findall('{http://purl.org/dc/elements/1.1/}title') if t.text]
                    titulo    = ' '.join(titulos) if titulos else ""
                    creadores = [c.text for c in dc.findall('{http://purl.org/dc/elements/1.1/}creator') if c.text]
                    descrips  = [d.text for d in dc.findall('{http://purl.org/dc/elements/1.1/}description') if d.text]
                    descripcion = ' '.join(descrips)[:500] if descrips else ""
                    fechas    = [f.text for f in dc.findall('{http://purl.org/dc/elements/1.1/}date') if f.text]
                    idents    = [i.text for i in dc.findall('{http://purl.org/dc/elements/1.1/}identifier') if i.text]
                    tipos     = [t.text.lower() for t in dc.findall('{http://purl.org/dc/elements/1.1/}type') if t.text]
                    subjects  = [s.text for s in dc.findall('{http://purl.org/dc/elements/1.1/}subject') if s.text]

                    año = None
                    for fecha in fechas:
                        m = re.search(r'(20\d{2})', str(fecha))
                        if m:
                            año = int(m.group(1)); break
                    if not año or año not in años_busqueda: continue

                    texto_full = f"{titulo} {descripcion} {' '.join(subjects)}"
                    relevancia = clasificar_relevancia(texto_full, var1, var2)
                    if relevancia == 'ninguna': continue

                    url = ""
                    for ident in idents:
                        if isinstance(ident, str) and ident.startswith('http'):
                            url = ident; break

                    es_tesis = any(t in ' '.join(tipos) for t in ['tesis','thesis','grado','maestr','doctor','bachelor','master','disser'])
                    if not es_tesis:
                        es_tesis = any(t in titulo.lower() for t in ['tesis','thesis'])

                    resultados.append({
                        "titulo": titulo, "autor": autores_display(creadores),
                        "autores_lista": creadores,
                        "cita_apa": formatear_cita_apa(creadores, año),
                        "resumen": descripcion, "año": año, "url": url, "citas": 0,
                        "tipo": "📘 Tesis" if es_tesis else "📄 Documento",
                        "fuente": f"OAI-PMH ({nombre_repo})",
                        "idioma": "🌎 OAI-PMH", "relevancia": relevancia
                    })
                except: continue
            token_elem = root.find('.//{http://www.openarchives.org/OAI/2.0/}resumptionToken')
            if token_elem is not None and token_elem.text and token_elem.text.strip():
                params = {'verb': 'ListRecords', 'resumptionToken': token_elem.text.strip()}
                pagina += 1; time.sleep(1)
            else: break
        except: break
    return resultados

# ════════════════════════════════════════════════════════════════════════════
# 🔬 APIs DE BÚSQUEDA
# ════════════════════════════════════════════════════════════════════════════
def buscar_openalex_articulos(var1, var2, año_inicio, año_fin, max_por_idioma=30):
    articulos, titulos_vistos = [], set()
    v1_en, v2_en = traducir(var1,'en'), (traducir(var2,'en') if var2 else '')
    v1_pt, v2_pt = traducir(var1,'pt'), (traducir(var2,'pt') if var2 else '')
    años_busqueda = list(range(año_inicio, año_fin+1))
    busquedas = {
        'es': [f"{var1} {var2}", var1] + ([var2] if var2 else []),
        'en': [f"{v1_en} {v2_en}", v1_en] + ([v2_en] if v2_en else []),
        'pt': [f"{v1_pt} {v2_pt}", v1_pt] + ([v2_pt] if v2_pt else []),
    }
    for idioma, queries in busquedas.items():
        for tema in queries:
            try:
                r = requests.get("https://api.openalex.org/works", params={
                    "search": tema, "per_page": max_por_idioma, "sort": "cited_by_count:desc",
                    "filter": f"type:article,language:{idioma},publication_year:{año_inicio}-{año_fin}",
                    "mailto": "investigador@universidad.edu"}, timeout=30)
                if r.status_code == 200:
                    for work in r.json().get("results",[]):
                        titulo = work.get("title","") or ""
                        if titulo in titulos_vistos: continue
                        abstract = get_abstract_openalex(work)
                        rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                        if rel == 'ninguna': continue
                        fecha = work.get("publication_date","")
                        año = int(fecha[:4]) if fecha and len(fecha)>=4 else 0
                        if año not in años_busqueda: continue
                        titulos_vistos.add(titulo)
                        autores = [a.get("author",{}).get("display_name","") for a in work.get("authorships",[]) if a.get("author",{}).get("display_name")]
                        articulos.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                            "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                            "url":work.get("doi") or work.get("id",""),"citas":safe_int(work.get("cited_by_count")),
                            "tipo":"📄 Artículo","fuente":f"OpenAlex ({IDIOMAS_LABEL.get(idioma,'')})","idioma":IDIOMAS_LABEL.get(idioma,''),
                            "relevancia":rel})
                time.sleep(0.3)
            except: pass
    return articulos

def buscar_openalex_tesis_latam(var1, var2, año_inicio, año_fin, max_results=30):
    tesis, titulos_vistos = [], set()
    tema = f"{var1} {var2}" if var2 else var1
    años_busqueda = list(range(año_inicio, año_fin+1))
    paises = ['PE','CO','MX','CL','EC','BR','AR','CR','BO','PY','UY','VE','PA','CU','DO','GT','HN','SV','NI']
    for pais_code in paises:
        try:
            r = requests.get("https://api.openalex.org/works", params={
                "search":tema,"per_page":max_results,"sort":"cited_by_count:desc",
                "filter":f"type:dissertation,institutions.country_code:{pais_code},publication_year:{año_inicio}-{año_fin}",
                "mailto":"investigador@universidad.edu"}, timeout=20)
            if r.status_code == 200:
                for work in r.json().get("results",[]):
                    titulo = work.get("title","") or ""
                    if titulo in titulos_vistos: continue
                    abstract = get_abstract_openalex(work)
                    rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                    if rel == 'ninguna': continue
                    fecha = work.get("publication_date","")
                    año = int(fecha[:4]) if fecha and len(fecha)>=4 else 0
                    if año not in años_busqueda: continue
                    titulos_vistos.add(titulo)
                    autores = [a.get("author",{}).get("display_name","") for a in work.get("authorships",[]) if a.get("author",{}).get("display_name")]
                    inst = ""
                    try:
                        auth = work.get("authorships",[])
                        if auth and auth[0].get("institutions"):
                            inst = auth[0]["institutions"][0].get("display_name","")
                    except: pass
                    autor_d = (autores[0] if autores else "") + (f" ({inst})" if inst else "")
                    tesis.append({"titulo":titulo,"autor":autor_d,"autores_lista":autores,
                        "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                        "url":work.get("doi") or work.get("id",""),"citas":safe_int(work.get("cited_by_count")),
                        "tipo":"📘 Tesis","fuente":f"OpenAlex (LATAM-{pais_code})","idioma":"🌎 LATAM",
                        "relevancia":rel})
            time.sleep(0.2)
        except: pass
    return tesis

def buscar_crossref(var1, var2, año_inicio, año_fin, max_results=30):
    resultados, titulos_vistos = [], set()
    años_busqueda = list(range(año_inicio, año_fin+1))
    query_en = f"{traducir(var1,'en')} {traducir(var2,'en')}" if var2 else traducir(var1,'en')
    queries = [query_en, traducir(var1,'en')] + ([traducir(var2,'en')] if var2 else [])
    for query in queries:
        try:
            r = requests.get("https://api.crossref.org/works", params={
                "query":query,"rows":max_results,"sort":"relevance",
                "filter":f"from-pub-date:{año_inicio},until-pub-date:{año_fin}",
                "select":"DOI,title,author,abstract,published-print,published-online,is-referenced-by-count,type"
            }, timeout=30, headers={"User-Agent":"AcademicSearchBot/2.0 (mailto:investigador@universidad.edu)"})
            if r.status_code == 200:
                for item in r.json().get("message",{}).get("items",[]):
                    titulo = item.get("title",[""])[0] if item.get("title") else ""
                    if titulo in titulos_vistos: continue
                    abstract = ""
                    if item.get("abstract"):
                        abstract = re.sub(r'<[^>]+>','', item.get("abstract",""))
                    rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                    if rel == 'ninguna': continue
                    autores_raw = item.get("author",[])
                    autores = [f"{a.get('given','')} {a.get('family','')}" .strip() for a in autores_raw]
                    fecha = item.get("published-print") or item.get("published-online") or {}
                    dp = fecha.get("date-parts",[[0]])[0]
                    año = dp[0] if dp else 0
                    if año not in años_busqueda: continue
                    titulos_vistos.add(titulo)
                    doi = item.get("DOI","")
                    resultados.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                        "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                        "url":f"https://doi.org/{doi}" if doi else "","citas":safe_int(item.get("is-referenced-by-count")),
                        "tipo":"📄 Artículo","fuente":"CrossRef","idioma":"🔗 CrossRef","relevancia":rel})
            time.sleep(0.3)
        except: pass
    return resultados

def buscar_pubmed(var1, var2, año_inicio, año_fin, max_results=30):
    resultados = []
    años_busqueda = list(range(año_inicio, año_fin+1))
    query_en = f"{traducir(var1,'en')} {traducir(var2,'en')}" if var2 else traducir(var1,'en')
    try:
        r = requests.get("https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi", params={
            "db":"pubmed","term":f"{query_en} AND {año_inicio}:{año_fin}[dp]",
            "retmax":max_results,"retmode":"json","sort":"relevance"}, timeout=30)
        if r.status_code != 200: return resultados
        ids = r.json().get("esearchresult",{}).get("idlist",[])
        if not ids: return resultados
        time.sleep(0.5)
        r = requests.get("https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi", params={
            "db":"pubmed","id":",".join(ids),"retmode":"xml"}, timeout=30)
        if r.status_code == 200:
            root = ET.fromstring(r.content)
            for article in root.findall('.//PubmedArticle'):
                titulo = article.findtext('.//ArticleTitle') or ""
                abstract = ' '.join([at.text or "" for at in article.findall('.//AbstractText')])
                rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                if rel == 'ninguna': continue
                autores = []
                for author in article.findall('.//Author'):
                    ln = author.findtext('LastName') or ""
                    fn = author.findtext('ForeName') or ""
                    if ln: autores.append(f"{fn} {ln}".strip())
                year = article.findtext('.//PubDate/Year') or ""
                m = re.search(r'(20\d{2})', str(year))
                año = int(m.group(1)) if m else 0
                if año not in años_busqueda: continue
                pmid = article.findtext('.//PMID') or ""
                resultados.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                    "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                    "url":f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else "","citas":0,
                    "tipo":"📄 Artículo","fuente":"PubMed","idioma":"🏥 PubMed","relevancia":rel})
    except: pass
    return resultados

def buscar_semantic_scholar(var1, var2, año_inicio, año_fin, max_results=30):
    resultados = []
    años_busqueda = list(range(año_inicio, año_fin+1))
    query_en = f"{traducir(var1,'en')} {traducir(var2,'en')}" if var2 else traducir(var1,'en')
    try:
        r = requests.get("https://api.semanticscholar.org/graph/v1/paper/search", params={
            "query":query_en,"limit":min(max_results,100),"year":f"{año_inicio}-{año_fin}",
            "fields":"title,authors,abstract,year,citationCount,externalIds,url"}, timeout=30)
        if r.status_code == 200:
            for paper in r.json().get("data",[]):
                titulo = paper.get("title","")
                abstract = paper.get("abstract","") or ""
                rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                if rel == 'ninguna': continue
                año = safe_int(paper.get("year"))
                if año not in años_busqueda: continue
                autores = [a.get("name","") for a in paper.get("authors",[]) if a.get("name")]
                ext = paper.get("externalIds",{}) or {}
                doi = ext.get("DOI","")
                url = f"https://doi.org/{doi}" if doi else paper.get("url","") or ""
                resultados.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                    "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                    "url":url,"citas":safe_int(paper.get("citationCount")),
                    "tipo":"📄 Artículo","fuente":"Semantic Scholar","idioma":"🤖 S.Scholar","relevancia":rel})
        elif r.status_code == 429: time.sleep(5)
    except: pass
    return resultados

def buscar_europe_pmc(var1, var2, año_inicio, año_fin, max_results=30):
    resultados = []
    años_busqueda = list(range(año_inicio, año_fin+1))
    query_en = f"{traducir(var1,'en')} {traducir(var2,'en')}" if var2 else traducir(var1,'en')
    try:
        r = requests.get("https://www.ebi.ac.uk/europepmc/webservices/rest/search", params={
            "query":f"{query_en} FIRST_PDATE:[{año_inicio} TO {año_fin}]",
            "format":"json","pageSize":max_results,"resultType":"core"}, timeout=30)
        if r.status_code == 200:
            for item in r.json().get("resultList",{}).get("result",[]):
                titulo = item.get("title","")
                abstract = item.get("abstractText","") or ""
                rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                if rel == 'ninguna': continue
                año = safe_int(item.get("pubYear"))
                if año not in años_busqueda: continue
                autores_str = item.get("authorString","")
                autores = [a.strip() for a in autores_str.split(",")[:10]] if autores_str else []
                doi  = item.get("doi","")
                pmid = item.get("pmid","")
                url  = f"https://doi.org/{doi}" if doi else (f"https://europepmc.org/article/MED/{pmid}" if pmid else "")
                resultados.append({"titulo":titulo,"autor":autores_str[:100],"autores_lista":autores,
                    "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                    "url":url,"citas":safe_int(item.get("citedByCount")),
                    "tipo":"📄 Artículo","fuente":"Europe PMC","idioma":"🧬 EuroPMC","relevancia":rel})
    except: pass
    return resultados

def buscar_doaj(var1, var2, año_inicio, año_fin, max_results=30):
    resultados = []
    años_busqueda = list(range(año_inicio, año_fin+1))
    query_en = f"{traducir(var1,'en')} {traducir(var2,'en')}" if var2 else traducir(var1,'en')
    try:
        r = requests.get(f"https://doaj.org/api/search/articles/{quote_plus(query_en)}",
                         params={"pageSize":max_results}, timeout=30)
        if r.status_code == 200:
            for item in r.json().get("results",[]):
                bib = item.get("bibjson",{})
                titulo = bib.get("title","")
                abstract = bib.get("abstract","") or ""
                rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                if rel == 'ninguna': continue
                año = safe_int(bib.get("year"))
                if año not in años_busqueda: continue
                autores = [a.get("name","") for a in bib.get("author",[]) if a.get("name")]
                links = bib.get("link",[])
                url = links[0].get("url","") if links else ""
                for ident in bib.get("identifier",[]):
                    if ident.get("type")=="doi":
                        url = f"https://doi.org/{ident.get('id','')}"; break
                resultados.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                    "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                    "url":url,"citas":0,
                    "tipo":"📄 Artículo","fuente":"DOAJ","idioma":"📖 DOAJ","relevancia":rel})
    except: pass
    return resultados

def buscar_arxiv(var1, var2, año_inicio, año_fin, max_results=30):
    resultados = []
    años_busqueda = list(range(año_inicio, año_fin+1))
    query_en = f"{traducir(var1,'en')} {traducir(var2,'en')}" if var2 else traducir(var1,'en')
    try:
        r = requests.get("http://export.arxiv.org/api/query", params={
            "search_query":f"all:{query_en}","max_results":max_results,
            "sortBy":"relevance","sortOrder":"descending"}, timeout=30)
        if r.status_code == 200:
            ns = {'atom':'http://www.w3.org/2005/Atom'}
            root = ET.fromstring(r.content)
            for entry in root.findall('atom:entry', ns):
                titulo  = (entry.findtext('atom:title','',ns) or "").replace('\n',' ').strip()
                abstract = (entry.findtext('atom:summary','',ns) or "").replace('\n',' ').strip()
                rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                if rel == 'ninguna': continue
                pub = entry.findtext('atom:published','',ns) or ""
                año = int(pub[:4]) if pub and len(pub)>=4 else 0
                if año not in años_busqueda: continue
                autores = [a.findtext('atom:name','',ns) for a in entry.findall('atom:author',ns)]
                url = entry.findtext('atom:id','',ns) or ""
                for link in entry.findall('atom:link',ns):
                    if link.get('title')=='pdf': url=link.get('href',url); break
                resultados.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                    "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                    "url":url,"citas":0,
                    "tipo":"📄 Preprint","fuente":"arXiv","idioma":"📄 arXiv","relevancia":rel})
    except: pass
    return resultados

def buscar_scielo(var1, var2, año_inicio, año_fin, max_results=30):
    resultados, titulos_vistos = [], set()
    años_busqueda = list(range(año_inicio, año_fin+1))
    tema_es = f"{var1} {var2}" if var2 else var1
    tema_pt = f"{traducir(var1,'pt')} {traducir(var2,'pt')}" if var2 else traducir(var1,'pt')
    for query in [tema_es, tema_pt]:
        try:
            r = requests.get("https://api.openalex.org/works", params={
                "search":query,"per_page":max_results,"sort":"cited_by_count:desc",
                "filter":f"primary_location.source.display_name.search:scielo,publication_year:{año_inicio}-{año_fin}",
                "mailto":"investigador@universidad.edu"}, timeout=30)
            if r.status_code == 200:
                for work in r.json().get("results",[]):
                    titulo = work.get("title","") or ""
                    if titulo in titulos_vistos: continue
                    abstract = get_abstract_openalex(work)
                    rel = clasificar_relevancia(f"{titulo} {abstract}", var1, var2)
                    if rel == 'ninguna': continue
                    fecha = work.get("publication_date","")
                    año = int(fecha[:4]) if fecha and len(fecha)>=4 else 0
                    if año not in años_busqueda: continue
                    titulos_vistos.add(titulo)
                    autores = [a.get("author",{}).get("display_name","") for a in work.get("authorships",[]) if a.get("author",{}).get("display_name")]
                    src = ""
                    try: src = work["primary_location"]["source"]["display_name"]
                    except: pass
                    resultados.append({"titulo":titulo,"autor":autores_display(autores),"autores_lista":autores,
                        "cita_apa":formatear_cita_apa(autores,año),"resumen":abstract[:500],"año":año,
                        "url":work.get("doi") or work.get("id",""),"citas":safe_int(work.get("cited_by_count")),
                        "tipo":"📄 Artículo","fuente":f"SciELO ({src[:25]})" if src else "SciELO","idioma":"🌎 SciELO",
                        "relevancia":rel})
            time.sleep(0.3)
        except: pass
    return resultados

# ════════════════════════════════════════════════════════════════════════════
# 🚀 EJECUTOR PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════
def ejecutar_busqueda(tema, año_inicio, año_fin, fuentes_seleccionadas, buscar_tesis_oai):
    var1, var2, pais = parsear_tema(tema)
    todos = []
    log_msgs = []

    MAPA_FUNCIONES = {
        "OpenAlex":         lambda: buscar_openalex_articulos(var1, var2, año_inicio, año_fin),
        "CrossRef":         lambda: buscar_crossref(var1, var2, año_inicio, año_fin),
        "PubMed":           lambda: buscar_pubmed(var1, var2, año_inicio, año_fin),
        "Semantic Scholar": lambda: buscar_semantic_scholar(var1, var2, año_inicio, año_fin),
        "Europe PMC":       lambda: buscar_europe_pmc(var1, var2, año_inicio, año_fin),
        "DOAJ":             lambda: buscar_doaj(var1, var2, año_inicio, año_fin),
        "arXiv":            lambda: buscar_arxiv(var1, var2, año_inicio, año_fin),
        "SciELO":           lambda: buscar_scielo(var1, var2, año_inicio, año_fin),
    }

    # Artículos
    for fuente in fuentes_seleccionadas:
        if fuente in MAPA_FUNCIONES:
            try:
                res = MAPA_FUNCIONES[fuente]()
                todos.extend(res)
                log_msgs.append(f"✅ {fuente}: {len(res)} resultados")
            except Exception as e:
                log_msgs.append(f"⚠️ {fuente}: error — {str(e)[:60]}")

    # Tesis OpenAlex LATAM (siempre)
    try:
        res = buscar_openalex_tesis_latam(var1, var2, año_inicio, año_fin)
        todos.extend(res)
        log_msgs.append(f"✅ Tesis OpenAlex LATAM: {len(res)} resultados")
    except Exception as e:
        log_msgs.append(f"⚠️ Tesis OA LATAM: {str(e)[:60]}")

    # OAI-PMH
    if buscar_tesis_oai:
        for nombre, url in REPOSITORIOS_OAIPMH.items():
            try:
                res = cosecha_oaipmh(url, nombre, var1, var2, año_inicio, año_fin)
                todos.extend(res)
                if res: log_msgs.append(f"✅ {nombre}: {len(res)} docs")
            except:
                pass
            time.sleep(0.3)

    # Deduplicar y ordenar
    if todos:
        df = pd.DataFrame(todos)
        df['_tn'] = df['titulo'].apply(normalizar)
        df = df.drop_duplicates(subset=['_tn'], keep='first').drop(columns=['_tn'])
        df['_ro'] = df['relevancia'].map({'ambas':0,'una':1})
        df = df.sort_values(['_ro','citas'], ascending=[True,False]).reset_index(drop=True)
        df = df.drop(columns=['_ro'])
        df['numero'] = range(1, len(df)+1)
        return df, var1, var2, pais, log_msgs
    return None, var1, var2, pais, log_msgs

# ════════════════════════════════════════════════════════════════════════════
# 📊 GRÁFICA CONSENSUS
# ════════════════════════════════════════════════════════════════════════════
def generar_grafica_consensus(df, tema, año_inicio, año_fin):
    random.seed(42)
    df = df.copy()

    df['y_pos'] = [random.uniform(0.53, 0.95) if r=='ambas' else random.uniform(0.05, 0.44)
                   for r in df['relevancia']]
    df['fecha_dt'] = [datetime(int(a), random.randint(1,12), random.randint(1,28)) for a in df['año']]

    max_citas = df['citas'].max() if df['citas'].max() > 0 else 1
    df['tamano'] = 22 + (df['citas'] / max_citas) * 34

    def color_punto(row):
        fuente, tipo = str(row.get('fuente','')), str(row.get('tipo',''))
        if 'Tesis' in tipo or 'Documento' in tipo:
            return '#f59e0b' if 'OAI-PMH' in fuente else '#8b5cf6'
        if 'OpenAlex' in fuente:
            if 'Inglés' in fuente:    return '#3b82f6'
            if 'Español' in fuente:   return '#ef4444'
            if 'Portugués' in fuente: return '#22c55e'
            if 'LATAM' in fuente:     return '#8b5cf6'
        if 'CrossRef'      in fuente: return '#06b6d4'
        if 'PubMed'        in fuente: return '#ec4899'
        if 'Semantic'      in fuente: return '#14b8a6'
        if 'Europe PMC'    in fuente: return '#6366f1'
        if 'DOAJ'          in fuente: return '#84cc16'
        if 'arXiv'         in fuente: return '#f97316'
        if 'SciELO'        in fuente: return '#10b981'
        return '#94a3b8'

    colores    = [color_punto(row) for _, row in df.iterrows()]
    bordes     = ['#ffffff' if r=='ambas' else '#4b5563' for r in df['relevancia']]
    ancho_bord = [3.5 if r=='ambas' else 2.0 for r in df['relevancia']]

    fig = go.Figure()

    años = list(range(año_inicio, año_fin+1))
    for año in años:
        fig.add_shape(type="line", x0=f"{año}-01-01", x1=f"{año}-01-01",
                      y0=0, y1=1, line=dict(color="#374151", width=1, dash="dot"))
        fig.add_annotation(x=f"{año}-06-15", y=0.98, text=f"<b>{año}</b>",
                          showarrow=False, font=dict(size=14, color='#6b7280'))

    fig.add_shape(type="line", x0=f"{año_inicio}-01-01", x1=f"{año_fin}-12-31",
                  y0=0.50, y1=0.50, line=dict(color="#22c55e", width=2, dash="dash"))

    fig.add_annotation(x=f"{año_inicio-1}-09-01", y=0.74,
                       text="<b>🎯 AMBAS VARIABLES</b>",
                       showarrow=False, font=dict(size=12, color='#22c55e'), xref='x', yref='y')
    fig.add_annotation(x=f"{año_inicio-1}-09-01", y=0.25,
                       text="<b>◎ UNA VARIABLE</b>",
                       showarrow=False, font=dict(size=12, color='#f59e0b'), xref='x', yref='y')

    titulos_h  = [str(t)[:65]+'…' if len(str(t))>65 else str(t) for t in df['titulo']]
    resumen_h  = [str(r)[:130]+'…' if r and len(str(r))>130 else (str(r) if r else '') for r in df['resumen']]
    etiqueta_h = ['🎯 Ambas variables' if r=='ambas' else '◎ Una variable' for r in df['relevancia']]

    fig.add_trace(go.Scatter(
        x=df['fecha_dt'], y=df['y_pos'],
        mode='markers+text',
        marker=dict(
            size=df['tamano'].tolist(),
            color=colores,
            line=dict(width=ancho_bord, color=bordes),
            opacity=0.93
        ),
        text=df['numero'],
        textfont=dict(color='white', size=10, family='Arial Black'),
        textposition='middle center',
        hovertemplate=(
            "<b style='font-size:13px'>%{customdata[7]}</b><br>"
            "<span style='font-size:11px;color:#94a3b8'>%{customdata[8]}</span><br><br>"
            "<b>%{customdata[0]}</b><br><br>"
            "<i style='font-size:11px'>%{customdata[6]}</i><br><br>"
            "%{customdata[4]} | %{customdata[5]}<br>"
            "📊 <b>%{customdata[2]} citas</b><br><br>"
            "<b style='color:#22c55e'>🖱️ CLIC PARA ABRIR</b><extra></extra>"
        ),
        customdata=list(zip(
            titulos_h, df['año'], df['citas'], df['url'],
            df['tipo'], df['fuente'], resumen_h, df['cita_apa'], etiqueta_h
        ))
    ))

    n_ambas_v = len(df[df['relevancia']=='ambas'])
    n_una_v   = len(df[df['relevancia']=='una'])
    n_art     = len(df[df['tipo'].str.contains('Artículo|Preprint')])
    n_tes     = len(df[df['tipo'].str.contains('Tesis|Documento')])

    fig.update_layout(
        title=dict(
            text=(f"<b>📚 Mapa de Investigación — Consensus</b><br>"
                  f"<span style='font-size:14px;color:#9ca3af'>{tema}</span><br>"
                  f"<span style='font-size:12px;color:#22c55e'>🖱️ Clic en círculo para abrir documento</span>"),
            font=dict(size=18, color='white'), x=0.5, y=0.96
        ),
        xaxis=dict(title="", showgrid=True, gridcolor='#1f2937', tickformat='%b %Y',
                   tickfont=dict(color='#9ca3af', size=10),
                   range=[f"{año_inicio-1}-06-01", f"{año_fin+1}-06-01"], fixedrange=True),
        yaxis=dict(visible=False, range=[0,1], fixedrange=True),
        plot_bgcolor='#111827', paper_bgcolor='#0f1923',
        font=dict(color='white'), showlegend=False, height=680,
        margin=dict(l=40, r=40, t=140, b=110), dragmode=False,
        annotations=[dict(
            text=("<b>ART:</b> 🔴ES 🔵EN 🟢PT 🩵CrossRef 🩷PubMed 🧊Semantic 🟣EuroPMC 🟡DOAJ 🟠arXiv 🌿SciELO"
                  "   │   <b>TESIS:</b> 🟠OAI-PMH 🟣OpenAlex   │   Tamaño = Citas   │   Borde blanco = ambas vars"),
            x=0.5, y=-0.07, xref='paper', yref='paper', showarrow=False, font=dict(size=9, color='#6b7280')
        )]
    )
    return fig, df

# ════════════════════════════════════════════════════════════════════════════
# 📥 EXPORTAR EXCEL (en memoria)
# ════════════════════════════════════════════════════════════════════════════
def generar_excel(df, tema, año_inicio, año_fin):
    df_excel = df[['numero','tipo','relevancia','cita_apa','titulo','autor','año','citas','resumen','fuente','url']].copy()
    df_excel['relevancia'] = df_excel['relevancia'].map({'ambas':'🎯 Ambas variables','una':'◎ Una variable'})
    df_excel.columns = ['N°','Tipo','Relevancia','Cita APA','Título','Autor(es)','Año','Citas','Resumen/Abstract','Fuente','URL']

    buf = io.BytesIO()
    anchos = [5, 15, 20, 26, 62, 36, 8, 8, 82, 28, 52]
    letras = 'ABCDEFGHIJK'

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_excel.to_excel(writer, sheet_name='Resultados', index=False)
        ws = writer.sheets['Resultados']

        header_fill   = PatternFill('solid', fgColor='1E3A5F')
        header_font   = Font(color='FFFFFF', bold=True, size=11)
        header_border = Border(
            bottom=Side(style='medium', color='0D47A1'),
            left=Side(style='thin', color='B0BEC5'),
            right=Side(style='thin', color='B0BEC5'),
            top=Side(style='thin', color='B0BEC5')
        )
        row_border = Border(
            bottom=Side(style='thin', color='E0E0E0'),
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0')
        )
        alt_fill = PatternFill('solid', fgColor='F1F8FF')

        for col_idx, (letra, ancho) in enumerate(zip(letras, anchos), 1):
            ws.column_dimensions[letra].width = ancho
            cell = ws.cell(row=1, column=col_idx)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28

        for row in range(2, len(df_excel)+2):
            ws.row_dimensions[row].height = 58
            for col_idx in range(1, 12):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = row_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = alt_fill

    buf.seek(0)
    nombre_limpio = re.sub(r'[^\w\s-]', '', tema)[:25].replace(' ','_')
    nombre_archivo = f"busqueda_{nombre_limpio}_{año_inicio}-{año_fin}.xlsx"
    return buf, nombre_archivo

# ════════════════════════════════════════════════════════════════════════════
# 🎨 PANTALLA DE LOGIN
# ════════════════════════════════════════════════════════════════════════════
def pantalla_login():
    st.markdown(CSS, unsafe_allow_html=True)
    # Centrar el login box
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div class="login-box">
            <h2 style="font-size:1.8rem; margin-bottom:6px;">🔐 Buscador Académico</h2>
            <p>Ingresa tus credenciales para continuar</p>
        </div>
        """, unsafe_allow_html=True)
        usuario  = st.text_input("👤 Usuario", placeholder="ej: admin", key="login_user")
        contrasena = st.text_input("🔑 Contraseña", type="password", placeholder="••••••••", key="login_pass")
        
        if st.button("Ingresar", key="btn_login"):
            if usuario in USUARIOS and USUARIOS[usuario] == contrasena:
                st.session_state["autenticado"] = True
                st.session_state["usuario_actual"] = usuario
                st.rerun()
            else:
                st.error("❌ Usuario o contraseña incorrectos.")

# ════════════════════════════════════════════════════════════════════════════
# 🎨 PANTALLA PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════
def pantalla_principal():
    st.markdown(CSS, unsafe_allow_html=True)

    # ─── Sidebar ──────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### 📚 Buscador Académico", unsafe_allow_html=False)
        st.markdown(f"Hola, **{st.session_state.get('usuario_actual','')}**")
        st.divider()

        st.markdown("**📝 Tema de búsqueda**")
        tema = st.text_input("", placeholder="ej: transformación digital y competitividad en Perú", key="input_tema")

        st.markdown("**📅 Rango de años**")
        col_a, col_b = st.columns(2)
        año_inicio = col_a.selectbox("Desde", list(range(2015,2026)), index=5, key="sel_año_ini")
        año_fin    = col_b.selectbox("Hasta", list(range(2015,2026)), index=10, key="sel_año_fin")

        st.markdown("**🔬 Fuentes de artículos**")
        fuentes_sel = st.multiselect("Selecciona", FUENTES_LISTA, default=FUENTES_LISTA, key="multi_fuentes")

        st.markdown("**🏛️ Opciones adicionales**")
        buscar_oai = st.checkbox("Buscar en repositorios OAI-PMH (20 repositorios LATAM)", value=True, key="chk_oai")

        st.divider()
        st.markdown("<span class='subtle'>📌 Tip: usa 'y' para separar dos variables.<br>Ej: <i>educación y tecnología en Colombia</i></span>", unsafe_allow_html=True)
        st.divider()

        if st.button("🚪 Cerrar sesión", key="btn_logout"):
            st.session_state.clear()
            st.rerun()

    # ─── Contenido principal ─────────────────────────────────────────────
    st.markdown("<h1>📚 Buscador Académico Multi-Fuente</h1>", unsafe_allow_html=True)
    st.markdown("<p class='centered subtle'>Busca tesis y artículos en 10+ fuentes científicas simultáneamente</p>", unsafe_allow_html=True)

    # Botón buscar (centrado, grande)
    col_c = st.columns([1,2,1])
    with col_c[1]:
        buscar_clicked = st.button("🔍  Buscar ahora", key="btn_buscar")

    # ─── Validaciones ────────────────────────────────────────────────────
    if buscar_clicked:
        if not tema.strip():
            st.warning("⚠️ Por favor ingresa un tema de búsqueda.")
            return
        if año_inicio > año_fin:
            st.warning("⚠️ El año de inicio debe ser menor o igual al año final.")
            return
        if not fuentes_sel:
            st.warning("⚠️ Selecciona al menos una fuente.")
            return

        # ─── Ejecutar búsqueda con progress ─────────────────────────────
        with st.spinner("🔄 Buscando en múltiples fuentes académicas... esto puede tomar un minuto."):
            df_res, var1, var2, pais, log = ejecutar_busqueda(tema, año_inicio, año_fin, fuentes_sel, buscar_oai)

        if df_res is None or df_res.empty:
            st.error("❌ No se encontraron documentos. Intenta con otros términos o amplía el rango de años.")
            with st.expander("📋 Log de búsqueda"):
                for msg in log: st.text(msg)
            return

        # ─── Guardar en session_state ────────────────────────────────────
        st.session_state["df_resultados"] = df_res
        st.session_state["tema_actual"]   = tema
        st.session_state["var1"]          = var1
        st.session_state["var2"]          = var2
        st.session_state["pais"]          = pais
        st.session_state["log"]           = log
        st.session_state["año_inicio"]    = año_inicio
        st.session_state["año_fin"]       = año_fin

    # ─── Mostrar resultados si existen ───────────────────────────────────
    if "df_resultados" in st.session_state:
        df_res     = st.session_state["df_resultados"]
        tema_act   = st.session_state["tema_actual"]
        var1       = st.session_state["var1"]
        var2       = st.session_state["var2"]
        año_inicio = st.session_state["año_inicio"]
        año_fin    = st.session_state["año_fin"]

        n_total  = len(df_res)
        n_ambas  = len(df_res[df_res['relevancia']=='ambas'])
        n_una    = len(df_res[df_res['relevancia']=='una'])
        n_art    = len(df_res[df_res['tipo'].str.contains('Artículo|Preprint')])
        n_tes    = len(df_res[df_res['tipo'].str.contains('Tesis|Documento')])
        n_fuentes= df_res['fuente'].nunique()

        # ── Stats cards ──────────────────────────────────────────────────
        st.markdown(f"""
        <div class="stats-row">
            <div class="stat-box"><div class="num">{n_total}</div><div class="lbl">📄 Total docs</div></div>
            <div class="stat-box"><div class="num" style="color:#22c55e">{n_ambas}</div><div class="lbl">🎯 Ambas vars</div></div>
            <div class="stat-box"><div class="num" style="color:#f59e0b">{n_una}</div><div class="lbl">◎ Una var</div></div>
            <div class="stat-box"><div class="num" style="color:#0ea5e9">{n_art}</div><div class="lbl">📄 Artículos</div></div>
            <div class="stat-box"><div class="num" style="color:#8b5cf6">{n_tes}</div><div class="lbl">📘 Tesis</div></div>
            <div class="stat-box"><div class="num" style="color:#f97316">{n_fuentes}</div><div class="lbl">📚 Fuentes</div></div>
        </div>
        """, unsafe_allow_html=True)

        # ── Variables detectadas ─────────────────────────────────────────
        st.markdown(f"""
        <div class="card" style="margin-bottom:14px; padding:14px 20px;">
            <span style="color:#7dd3fc; font-weight:600;">🔬 Variable 1:</span>
            <span style="color:#e2e8f0; margin-left:8px;">{var1}</span>
            {"<br><span style='color:#7dd3fc; font-weight:600;'>🔬 Variable 2:</span> <span style='color:#e2e8f0; margin-left:8px;'>" + str(var2) + "</span>" if var2 else ""}
            {"<br><span style='color:#7dd3fc; font-weight:600;'>🌍 País:</span> <span style='color:#e2e8f0; margin-left:8px;'>" + str(st.session_state.get("pais","")) + "</span>" if st.session_state.get("pais") else ""}
        </div>
        """, unsafe_allow_html=True)

        # ── Gráfica Consensus ────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📊 Mapa de Investigación (Consensus)", unsafe_allow_html=False)
        fig, df_grafica = generar_grafica_consensus(df_res, tema_act, año_inicio, año_fin)

        # Generar HTML con click-to-open
        fig_json  = fig.to_json()
        urls_json = json.dumps(df_grafica['url'].tolist())
        html_chart = f"""
        <div id="consensus_chart" style="width:100%; height:680px; border-radius:12px; background:#111827;"></div>
        <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
        <script>
            var fig  = {fig_json};
            var urls = {urls_json};
            var div  = document.getElementById("consensus_chart");
            Plotly.newPlot(div, fig.data, fig.layout, {{displayModeBar:false, scrollZoom:false}});
            div.on("plotly_click", function(d) {{
                var u = urls[d.points[0].pointIndex];
                if (u) window.open(u, "_blank");
            }});
        </script>
        """
        st.components.v1.html(html_chart, height=720, scrolling=False)

        # ── Botones de descarga ──────────────────────────────────────────
        st.markdown("---")
        col_dl = st.columns([1,1,1])

        # Excel
        with col_dl[0]:
            buf_xlsx, nombre_xlsx = generar_excel(df_res, tema_act, año_inicio, año_fin)
            st.download_button(
                label="📥 Descargar Excel",
                data=buf_xlsx,
                file_name=nombre_xlsx,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel"
            )

        # CSV
        with col_dl[1]:
            cols_csv = ['numero','tipo','relevancia','cita_apa','titulo','autor','año','citas','resumen','fuente','url']
            csv_data = df_res[cols_csv].to_csv(index=False, encoding='utf-8')
            nombre_csv = re.sub(r'[^\w\s-]', '', tema_act)[:25].replace(' ','_')
            st.download_button(
                label="📥 Descargar CSV",
                data=csv_data.encode('utf-8'),
                file_name=f"busqueda_{nombre_csv}_{año_inicio}-{año_fin}.csv",
                mime="text/csv",
                key="dl_csv"
            )

        # Log
        with col_dl[2]:
            log_text = "\n".join(st.session_state.get("log",[]))
            st.download_button(
                label="📋 Descargar Log",
                data=log_text.encode('utf-8'),
                file_name="log_busqueda.txt",
                mime="text/plain",
                key="dl_log"
            )

        # ── Tabla de resultados ──────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Lista de Documentos", unsafe_allow_html=False)

        # Filtros rápidos
        col_f1, col_f2 = st.columns(2)
        filtro_tipo = col_f1.selectbox("Filtrar por tipo", ["Todos","📄 Artículos/Preprints","📘 Tesis/Documentos"], key="sel_tipo")
        filtro_rel  = col_f2.selectbox("Filtrar por relevancia", ["Todos","🎯 Ambas variables","◎ Una variable"], key="sel_rel")

        df_filtrado = df_res.copy()
        if filtro_tipo == "📄 Artículos/Preprints":
            df_filtrado = df_filtrado[df_filtrado['tipo'].str.contains('Artículo|Preprint')]
        elif filtro_tipo == "📘 Tesis/Documentos":
            df_filtrado = df_filtrado[df_filtrado['tipo'].str.contains('Tesis|Documento')]
        if filtro_rel == "🎯 Ambas variables":
            df_filtrado = df_filtrado[df_filtrado['relevancia']=='ambas']
        elif filtro_rel == "◎ Una variable":
            df_filtrado = df_filtrado[df_filtrado['relevancia']=='una']

        for _, row in df_filtrado.iterrows():
            badge_rel = '<span class="badge-ambas">🎯 Ambas variables</span>' if row['relevancia']=='ambas' else '<span class="badge-una">◎ Una variable</span>'
            badge_tipo = '<span class="badge-tesis">📘 Tesis</span>' if 'Tesis' in str(row['tipo']) or 'Documento' in str(row['tipo']) else '<span class="badge-articulo">📄 Artículo</span>'
            url_link = f'<a href="{row["url"]}" target="_blank" style="color:#38bdf8;">🔗 Ver documento</a>' if row.get('url') else ''

            resumen_short = str(row.get('resumen',''))[:160] + '…' if row.get('resumen') and len(str(row.get('resumen','')))>160 else str(row.get('resumen',''))

            st.markdown(f"""
            <div class="result-card">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                    <span style="color:#64748b; font-weight:700; font-size:0.85rem;">#{row['numero']}</span>
                    <div>{badge_rel} &nbsp; {badge_tipo}</div>
                </div>
                <div style="font-size:1rem; font-weight:600; color:#e2e8f0; margin-bottom:5px;">{row['titulo']}</div>
                <div style="font-size:0.82rem; color:#94a3b8; margin-bottom:4px;">
                    <b style="color:#7dd3fc;">{row['cita_apa']}</b> &nbsp;|&nbsp; {row['autor']} &nbsp;|&nbsp; 📊 {row['citas']} citas &nbsp;|&nbsp; {row['fuente']}
                </div>
                <div style="font-size:0.79rem; color:#64748b; margin-bottom:6px;">{resumen_short}</div>
                <div>{url_link}</div>
            </div>
            """, unsafe_allow_html=True)

        # Log desplegable
        with st.expander("📋 Log de búsqueda"):
            for msg in st.session_state.get("log",[]): st.text(msg)

# ════════════════════════════════════════════════════════════════════════════
# 🏁 MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="📚 Buscador Académico Multi-Fuente",
        page_icon="📚",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    if not st.session_state.get("autenticado", False):
        pantalla_login()
    else:
        pantalla_principal()

if __name__ == "__main__":
    main()
